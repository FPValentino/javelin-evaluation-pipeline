"""
Bug Type Classification & Performance Analysis Script
=====================================================
Classifies each Defects4J bug by structural type (from patch analysis)
and cross-tabulates with Ochiai vs Ochiai-MS performance.

Inputs:
  - All Patches/          (136 patch files)
  - root_causes.json      (Defects4J official root causes)
  - ground_truth.json     (fault locations)
  - Javelin_Evaluation_Report_avg.xlsx   (average ranking results)
  - Javelin_Evaluation_Report_dense.xlsx (dense ranking results)

Output:
  - Bug_Type_Analysis.xlsx
"""

import os
import re
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter


BUG_TYPE_DESCRIPTIONS = {
    "Wrong Condition": (
        "A conditional expression (if, while, ternary, or logical operator) "
        "evaluates incorrectly. The fix modifies the condition's logic "
        "(e.g., != to ==, && to ||, < to <=)."
    ),
    "Missing Condition": (
        "A necessary conditional check is absent. The fix adds a new if-statement, "
        "else-branch, or guard clause that was not present in the buggy version."
    ),
    "Wrong Value/Variable": (
        "An incorrect constant, variable, return value, or expression is used. "
        "The fix replaces one value/variable with another without changing "
        "control flow structure."
    ),
    "Missing Method Call": (
        "A required method invocation is absent. The fix adds a call to a method "
        "that should have been invoked (e.g., adding a close(), flush(), or "
        "initialization call)."
    ),
    "Wrong Method Call": (
        "The wrong method is called, or correct method is called with wrong arguments. "
        "The fix changes the method name or its parameters."
    ),
    "Null Handling": (
        "A null pointer dereference or missing null check. The fix adds a null guard, "
        "changes null-related logic, or adds a null-safe alternative."
    ),
    "Type Error": (
        "Incorrect type casting, generic type usage, or instanceof check. "
        "The fix changes type-related operations (casts, type parameters, "
        "instanceof conditions)."
    ),
    "Missing Code": (
        "A significant block of logic is entirely absent (more than a simple condition "
        "or single statement). The fix adds multiple lines of new functionality "
        "such as a new method, a new handling path, or a substantial logic block."
    ),
    "Exception Handling": (
        "Incorrect or missing try-catch-throw-finally logic. The fix adds, removes, "
        "or modifies exception handling code."
    ),
    "Data Structure": (
        "Wrong collection type, incorrect data structure operation, or wrong "
        "initialization of arrays/lists/maps. The fix changes how data is stored "
        "or accessed in collections (e.g., HashMap to HashSet)."
    ),
    "Algorithm/Logic": (
        "The computational logic or algorithm is incorrect. The fix changes "
        "arithmetic operations, loop bounds, algorithmic steps, or refactors "
        "the approach to computation."
    ),
}


# --- PATCH CLASSIFICATION HEURISTICS (refined) ---

def classify_patch(patch_text):
    """Classify a bug based on its patch diff content.

    Priority order:
      1. Missing Code       (large additions, new methods)
      2. Data Structure     (collection type changes)
      3. Exception Handling (try/catch/throw dominant)
      4. Null Handling      (null is PRIMARY change, strict)
      5. Wrong Condition    (condition logic modified)
      6. Missing Condition  (new guard/if added)
      7. Type Error         (cast/instanceof changes)
      8. Wrong Method Call  (different method called)
      9. Missing Method Call(new call added)
     10. Wrong Value/Variable (simple swap)
     11. Algorithm/Logic    (complex, fallback)
    """
    added_lines = []
    removed_lines = []

    for line in patch_text.split("\n"):
        if line.startswith("+") and not line.startswith("+++"):
            content = line[1:].strip()
            if content and not content.startswith("//"):
                added_lines.append(content)
        elif line.startswith("-") and not line.startswith("---"):
            content = line[1:].strip()
            if content and not content.startswith("//"):
                removed_lines.append(content)

    added_text = "\n".join(added_lines)
    removed_text = "\n".join(removed_lines)

    num_added = len(added_lines)
    num_removed = len(removed_lines)

    # --- 1. MISSING CODE: large pure additions or new method definitions ---
    if num_added > 10 and num_removed <= 3:
        return "Missing Code"
    # New method defined in added code (not in removed)
    new_method_pattern = r'(public|private|protected|static)\s+.*\w+\s*\([^)]*\)\s*\{?'
    added_methods = re.findall(new_method_pattern, added_text)
    removed_methods = re.findall(new_method_pattern, removed_text)
    if len(added_methods) > len(removed_methods) and num_added > 8:
        return "Missing Code"

    # --- 2. DATA STRUCTURE: collection type swap or structural data change ---
    collection_types = [
        "HashMap", "HashSet", "TreeMap", "TreeSet", "LinkedHashMap",
        "LinkedHashSet", "ArrayList", "LinkedList", "ConcurrentHashMap",
        "Vector", "Hashtable", "ArrayDeque",
    ]
    abstract_types = ["Map", "Set", "List", "Queue", "Deque", "Collection"]
    removed_collections = [t for t in collection_types + abstract_types if t in removed_text]
    added_collections = [t for t in collection_types + abstract_types if t in added_text]
    if removed_collections and added_collections:
        # Different collection types between removed and added
        if set(removed_collections) != set(added_collections):
            return "Data Structure"
    # Array to collection or vice versa
    if (re.search(r'new\s+\w+\[', removed_text) and re.search(r'new\s+(Array|Linked|Hash)', added_text)):
        return "Data Structure"
    if (re.search(r'new\s+(Array|Linked|Hash)', removed_text) and re.search(r'new\s+\w+\[', added_text)):
        return "Data Structure"

    # --- 3. EXCEPTION HANDLING: try/catch/throw dominant ---
    exc_keywords = [r'\btry\b', r'\bcatch\b', r'\bthrow\s', r'\bthrows\b', r'\bfinally\b']
    exc_added = sum(1 for l in added_lines if any(re.search(p, l) for p in exc_keywords))
    exc_removed = sum(1 for l in removed_lines if any(re.search(p, l) for p in exc_keywords))
    if (exc_added + exc_removed) >= 2:
        total_lines = num_added + num_removed
        if (exc_added + exc_removed) / max(total_lines, 1) > 0.25:
            return "Exception Handling"
    # Pure addition of throw/try-catch
    if exc_added >= 2 and num_removed == 0:
        return "Exception Handling"

    # --- 4. NULL HANDLING: null is the PRIMARY change (strict) ---
    null_added = sum(1 for l in added_lines if re.search(r'\bnull\b', l))
    null_removed = sum(1 for l in removed_lines if re.search(r'\bnull\b', l))
    # If BOTH removed and added have null-in-condition, it's a condition change not null handling
    removed_has_null_cond = bool(re.search(r'\bif\b.*\bnull\b', removed_text))
    added_has_null_cond = bool(re.search(r'\bif\b.*\bnull\b', added_text))
    # Strict: majority of non-trivial added lines must involve null
    if null_added > 0 and num_added > 0 and not removed_has_null_cond:
        null_ratio = null_added / num_added
        if null_ratio >= 0.4 and added_has_null_cond:
            if num_added <= 8:
                return "Null Handling"
    # Simple null-safe ternary: x == null ? y : z
    if num_added <= 3 and num_removed <= 3:
        if re.search(r'==\s*null\s*\?', added_text) or re.search(r'!=\s*null\s*\?', added_text):
            return "Null Handling"
    # Adding a null guard where there wasn't one
    if num_removed == 0 and num_added <= 5:
        if re.search(r'\bif\s*\(.*==\s*null', added_text) or re.search(r'\bif\s*\(.*!=\s*null', added_text):
            return "Null Handling"

    # --- 5. WRONG CONDITION: existing condition logic is changed ---
    if num_removed > 0 and num_added > 0:
        removed_conds = [l for l in removed_lines if re.search(r'\bif\s*\(|\bwhile\s*\(|\bfor\s*\(', l)]
        added_conds = [l for l in added_lines if re.search(r'\bif\s*\(|\bwhile\s*\(|\bfor\s*\(', l)]
        if removed_conds and added_conds and len(removed_conds) >= len(added_conds):
            # Operators changed in condition
            operators = [r'==', r'!=', r'<=', r'>=', r'<(?!=)', r'>(?!=)', r'&&', r'\|\|', r'\bnull\b']
            for op in operators:
                r_has = any(re.search(op, l) for l in removed_conds)
                a_has = any(re.search(op, l) for l in added_conds)
                if r_has != a_has or (r_has and a_has):
                    return "Wrong Condition"
            # Condition expression itself changed (same structure, different content)
            if len(removed_conds) == len(added_conds):
                return "Wrong Condition"

    # --- 6. MISSING CONDITION: new if/guard added where none existed ---
    added_ifs = len(re.findall(r'\bif\s*\(', added_text))
    removed_ifs = len(re.findall(r'\bif\s*\(', removed_text))
    if added_ifs > removed_ifs:
        return "Missing Condition"

    # --- 7. TYPE ERROR: casting, instanceof, generics ---
    type_patterns = [r'\binstanceof\b', r'\(\w+(<[^>]*>)?\)\s*\w', r'\bClass<', r'\.class\b']
    if num_removed > 0 and num_added > 0:
        type_removed = sum(1 for l in removed_lines if any(re.search(p, l) for p in type_patterns))
        type_added = sum(1 for l in added_lines if any(re.search(p, l) for p in type_patterns))
        if type_removed > 0 and type_added > 0:
            type_ratio = (type_removed + type_added) / max(num_added + num_removed, 1)
            if type_ratio > 0.3:
                return "Type Error"

    # --- 8. WRONG METHOD CALL: method name or arguments changed ---
    if num_removed > 0 and num_added > 0 and num_removed <= 5 and num_added <= 5:
        removed_calls = re.findall(r'\.(\w+)\s*\(', removed_text)
        added_calls = re.findall(r'\.(\w+)\s*\(', added_text)
        if removed_calls and added_calls:
            removed_set = set(removed_calls)
            added_set = set(added_calls)
            # Methods that appear in removed but not added (replaced)
            if removed_set - added_set and added_set - removed_set:
                return "Wrong Method Call"

    # --- 9. MISSING METHOD CALL: new call added ---
    if num_removed == 0 and num_added > 0 and num_added <= 4:
        if re.search(r'\.\w+\s*\(', added_text) and not re.search(r'\bif\s*\(', added_text):
            return "Missing Method Call"

    # --- 10. WRONG VALUE/VARIABLE: simple value swap ---
    if num_removed > 0 and num_added > 0 and num_removed <= 4 and num_added <= 4:
        if not re.search(r'\bif\b|\bwhile\b|\bfor\b|\btry\b|\bcatch\b', added_text + removed_text):
            return "Wrong Value/Variable"
        # Even with some control flow, if it's a small change it's likely a value fix
        if num_removed <= 2 and num_added <= 2:
            return "Wrong Value/Variable"

    # --- 11. ALGORITHM/LOGIC: complex changes that don't fit above ---
    if num_added > 3 or num_removed > 3:
        return "Algorithm/Logic"

    # Fallback
    if num_added > 0 and num_removed > 0:
        return "Wrong Value/Variable"
    elif num_added > 0:
        return "Missing Code"
    else:
        return "Algorithm/Logic"


# --- DATA LOADING ---

def load_evaluation_data(xlsx_path):
    """Load bug-level EXAM scores and ranks from evaluation xlsx."""
    df = pd.read_excel(xlsx_path, sheet_name="Bug-Level Scores", header=3)
    results = {}
    for _, row in df.iterrows():
        bug_id = row.get("Bug ID")
        if pd.isna(bug_id) or not str(bug_id).startswith("Defects4J"):
            continue
        results[str(bug_id)] = {
            "ochiai_exam": row.get("Ochiai EXAM"),
            "ochiai_ms_exam": row.get("Ochiai-MS EXAM"),
            "ochiai_rank": row.get("Ochiai Best Rank"),
            "ochiai_ms_rank": row.get("Ochiai-MS Best Rank"),
            "ochiai_gt": row.get("Ochiai GT Found"),
            "ochiai_ms_gt": row.get("Ochiai-MS GT Found"),
        }
    return results


def determine_winner(ochiai_rank, ochiai_ms_rank):
    """Compare ranks. Lower rank = found fault sooner = better."""
    if pd.isna(ochiai_rank) and pd.isna(ochiai_ms_rank):
        return "Both Undetected"
    if pd.isna(ochiai_rank):
        return "Ochiai-MS Better"
    if pd.isna(ochiai_ms_rank):
        return "Ochiai Better"

    ochiai_rank = float(ochiai_rank)
    ochiai_ms_rank = float(ochiai_ms_rank)

    if ochiai_ms_rank < ochiai_rank:
        return "Ochiai-MS Better"
    elif ochiai_ms_rank > ochiai_rank:
        return "Ochiai Better"
    else:
        return "Tie"


def blank_if_zero(val):
    """Return empty string if value is 0, else return the value."""
    if val == 0:
        return ""
    return val


# --- MAIN ---

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    patches_dir = os.path.join(base_dir, "All Patches")

    # Load data
    print("Loading data...")
    with open(os.path.join(base_dir, "ground_truth.json"), "r") as f:
        ground_truth = json.load(f)

    with open(os.path.join(base_dir, "root_causes.json"), "r") as f:
        root_causes = json.load(f)

    avg_data = load_evaluation_data(os.path.join(base_dir, "Javelin_Evaluation_Report_avg.xlsx"))
    dense_data = load_evaluation_data(os.path.join(base_dir, "Javelin_Evaluation_Report_dense.xlsx"))

    # Classify each bug
    print("Classifying bugs from patches...")
    bug_records = []

    for bug_id in sorted(ground_truth.keys()):
        parts = bug_id.replace("Defects4J-", "").rsplit("-", 1)
        project = parts[0]
        bug_num = parts[1]

        patch_path = os.path.join(patches_dir, f"{bug_id}.patch")
        if os.path.exists(patch_path):
            with open(patch_path, "r", encoding="utf-8", errors="ignore") as f:
                patch_text = f.read()
            bug_type = classify_patch(patch_text)
        else:
            bug_type = "Unknown"
            print(f"  WARNING: No patch file for {bug_id}")

        root_cause = root_causes.get(bug_id, "N/A")
        if root_cause.startswith("-> "):
            root_cause = root_cause[3:]

        avg_info = avg_data.get(bug_id, {})
        dense_info = dense_data.get(bug_id, {})

        ochiai_rank_avg = avg_info.get("ochiai_rank")
        ochiai_ms_rank_avg = avg_info.get("ochiai_ms_rank")
        ochiai_exam_avg = avg_info.get("ochiai_exam")
        ochiai_ms_exam_avg = avg_info.get("ochiai_ms_exam")

        ochiai_rank_dense = dense_info.get("ochiai_rank")
        ochiai_ms_rank_dense = dense_info.get("ochiai_ms_rank")
        ochiai_exam_dense = dense_info.get("ochiai_exam")
        ochiai_ms_exam_dense = dense_info.get("ochiai_ms_exam")

        winner_avg = determine_winner(ochiai_rank_avg, ochiai_ms_rank_avg)
        winner_dense = determine_winner(ochiai_rank_dense, ochiai_ms_rank_dense)

        bug_records.append({
            "Bug ID": bug_id,
            "Project": project,
            "Bug #": bug_num,
            "Bug Type": bug_type,
            "Root Cause (Defects4J)": root_cause,
            "Ochiai Rank (Avg)": ochiai_rank_avg,
            "Ochiai-MS Rank (Avg)": ochiai_ms_rank_avg,
            "Ochiai EXAM (Avg)": ochiai_exam_avg,
            "Ochiai-MS EXAM (Avg)": ochiai_ms_exam_avg,
            "Winner (Avg Ranking)": winner_avg,
            "Ochiai Rank (Dense)": ochiai_rank_dense,
            "Ochiai-MS Rank (Dense)": ochiai_ms_rank_dense,
            "Ochiai EXAM (Dense)": ochiai_exam_dense,
            "Ochiai-MS EXAM (Dense)": ochiai_ms_exam_dense,
            "Winner (Dense Ranking)": winner_dense,
        })

    print(f"Classified {len(bug_records)} bugs.")

    # --- BUILD EXCEL OUTPUT ---
    print("Building Excel report...")
    wb = Workbook()

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== SHEET 1: GUIDE =====
    ws = wb.active
    ws.title = "Guide"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 105

    ws.append(["Topic", "Description"])
    ws['A1'].font = header_font
    ws['B1'].font = header_font

    ws.append(["REPORT OVERVIEW",
               "This report classifies each of the 136 Defects4J bugs by structural type "
               "(based on patch diff analysis) and cross-tabulates with Ochiai vs Ochiai-MS "
               "fault localization performance."])
    ws.append(["",
               "The goal is to identify which types of bugs Ochiai-MS performs better, worse, "
               "or equally compared to Standard Ochiai."])
    ws.append(["", ""])
    ws.append(["--- METHODOLOGY ---", ""])
    ws.append(["Bug Classification",
               "Each bug is classified by analyzing its patch file (diff between buggy and "
               "fixed versions). The structural pattern of the code change determines the type. "
               "Taxonomy based on Sobreira et al. (2018) - Dissection of a Bug Dataset."])
    ws.append(["Performance Comparison",
               "Winner is determined by comparing Best Rank (lines inspected to find first fault). "
               "Lower rank = found the fault sooner = better performance. "
               "Tie = both algorithms rank the fault at the exact same position."])
    ws.append(["Root Cause",
               "The official Defects4J root cause from the triggering test (exception/error type). "
               "Provided as supplementary context - not used for classification."])
    ws.append(["", ""])
    ws.append(["--- BUG TYPE TAXONOMY ---", ""])
    ws.append(["Category", "Definition"])
    ws['A11'].font = header_font
    ws['B11'].font = header_font

    for bug_type, desc in BUG_TYPE_DESCRIPTIONS.items():
        ws.append([bug_type, desc])

    ws.append(["", ""])
    ws.append(["--- CLASSIFICATION METHOD ---", ""])
    ws.append(["How It Works",
               "The script reads each patch file (diff between buggy and fixed code) and "
               "examines the added (+) and removed (-) lines. It applies the following rules "
               "in priority order (first match wins):"])
    ws.append(["1. Missing Code",
               "If >10 lines are added with <=3 removed, or a new method definition appears "
               "in the additions. The bug was about entirely absent logic."])
    ws.append(["2. Data Structure",
               "If the removed lines contain one collection type (e.g., HashMap, Map) and "
               "the added lines contain a different type (e.g., HashSet, Set). The fix "
               "swapped or restructured how data is stored."])
    ws.append(["3. Exception Handling",
               "If >25% of changed lines involve try/catch/throw/finally keywords."])
    ws.append(["4. Null Handling",
               "If the majority (>=40%) of added lines involve 'null' AND a new "
               "if(...null...) guard is added that did NOT exist in the removed code. "
               "Strict: does not fire when the removed code already had a null condition "
               "(that case is a Wrong Condition instead)."])
    ws.append(["5. Wrong Condition",
               "If BOTH the removed and added code contain if/while/for conditions "
               "(i.e., an existing condition was modified). Detects operator changes "
               "(!=, ==, <, >, &&, ||) and expression modifications."])
    ws.append(["6. Missing Condition",
               "If the added code has MORE if-statements than the removed code, meaning "
               "a new guard or conditional branch was introduced."])
    ws.append(["7. Type Error",
               "If >30% of changes involve instanceof, type casts, or generic types."])
    ws.append(["8. Wrong Method Call",
               "If the removed code calls method X and the added code calls a different "
               "method Y (small patch, <=5 lines)."])
    ws.append(["9. Missing Method Call",
               "If a new method call is added with no removals and no control flow change."])
    ws.append(["10. Wrong Value/Variable",
               "Small changes (<=4 lines each) that do not involve control flow keywords. "
               "Likely a constant, variable name, or return value swap."])
    ws.append(["11. Algorithm/Logic",
               "Fallback for complex changes that do not match any of the above patterns."])
    ws.append(["", ""])
    ws.append(["--- RANKING METHODS ---", ""])
    ws.append(["Average (Midpoint) Ranking",
               "Tied suspiciousness scores receive the midpoint of their ordinal positions. "
               "E.g., 3 lines tied at positions 5-7 all get rank 6.0. "
               "This is the primary ranking method used in SBFL literature."])
    ws.append(["Dense Ranking",
               "Consecutive integer ranks where ties share the same rank number. "
               "E.g., if 3 lines are tied at rank 2, the next distinct score is rank 3. "
               "Lines inspected = count of lines at or above the fault's rank."])
    ws.append(["", ""])
    ws.append(["--- NOTES ---", ""])
    ws.append(["High Ranks (e.g., >20000)",
               "Some bugs have very high ranks because the faulty line receives a "
               "suspiciousness score of 0 (tied with thousands of other lines). This "
               "is common for omission bugs where the fix ADDS code that never existed - "
               "since the location was never differentially executed, SBFL cannot "
               "distinguish it from background. This is a known limitation of "
               "spectrum-based fault localization for omission/missing-code bugs."])

    # ===== SHEET 2: BUG-LEVEL CLASSIFICATION =====
    ws2 = wb.create_sheet("Bug-Level Classification")

    ws2.append(["BUG-LEVEL CLASSIFICATION: Each bug with its structural type, root cause, and performance comparison."])
    ws2.append(["Winner = algorithm that ranks the fault higher (fewer lines to inspect). Tie = same rank position."])
    ws2.append([""])

    headers = ["Bug ID", "Project", "Bug #", "Bug Type", "Root Cause (Defects4J)",
               "Ochiai Rank (Avg)", "Ochiai-MS Rank (Avg)", "Winner (Avg)",
               "Ochiai EXAM (Avg)", "Ochiai-MS EXAM (Avg)",
               "Ochiai Rank (Dense)", "Ochiai-MS Rank (Dense)", "Winner (Dense)",
               "Ochiai EXAM (Dense)", "Ochiai-MS EXAM (Dense)"]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        ws2.cell(row=4, column=col).font = header_font
        ws2.cell(row=4, column=col).border = thin_border

    project_order = ["Chart", "Cli", "Csv", "Gson", "JacksonCore", "JacksonDatabind", "Jsoup", "Lang"]
    bug_records.sort(key=lambda x: (project_order.index(x["Project"]) if x["Project"] in project_order else 99, int(x["Bug #"])))

    for rec in bug_records:
        row_data = [
            rec["Bug ID"], rec["Project"], rec["Bug #"], rec["Bug Type"],
            rec["Root Cause (Defects4J)"],
            rec["Ochiai Rank (Avg)"], rec["Ochiai-MS Rank (Avg)"], rec["Winner (Avg Ranking)"],
            rec["Ochiai EXAM (Avg)"], rec["Ochiai-MS EXAM (Avg)"],
            rec["Ochiai Rank (Dense)"], rec["Ochiai-MS Rank (Dense)"], rec["Winner (Dense Ranking)"],
            rec["Ochiai EXAM (Dense)"], rec["Ochiai-MS EXAM (Dense)"],
        ]
        ws2.append(row_data)
        row_num = ws2.max_row
        for col_idx in [8, 13]:
            cell = ws2.cell(row=row_num, column=col_idx)
            if cell.value == "Ochiai-MS Better":
                cell.fill = green_fill
            elif cell.value == "Ochiai Better":
                cell.fill = red_fill
            elif cell.value == "Tie":
                cell.fill = yellow_fill
            elif cell.value == "Both Undetected":
                cell.fill = gray_fill

    col_widths = [22, 18, 6, 22, 60, 18, 20, 18, 18, 20, 18, 20, 18, 18, 20]
    for col, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ===== SHEET 3: BY BUG TYPE (Cross-tabulation) =====
    ws3 = wb.create_sheet("By Bug Type")

    ws3.append(["CROSS-TABULATION: Bug Type x Performance Outcome"])
    ws3.append(["Shows how many bugs of each type Ochiai-MS handles better, worse, or ties with Ochiai."])
    ws3.append(["Percentages exclude 'Both Undetected' bugs (only detectable bugs considered)."])
    ws3.append([""])

    # --- Average Ranking Section ---
    ws3.append(["AVERAGE (MIDPOINT) RANKING"])
    ws3.cell(row=ws3.max_row, column=1).font = title_font
    ws3.append([""])

    type_perf_avg = defaultdict(lambda: {"Ochiai-MS Better": 0, "Ochiai Better": 0, "Tie": 0, "Both Undetected": 0, "Total": 0})
    for rec in bug_records:
        bt = rec["Bug Type"]
        winner = rec["Winner (Avg Ranking)"]
        type_perf_avg[bt][winner] += 1
        type_perf_avg[bt]["Total"] += 1

    headers3 = ["Bug Type", "Total Bugs", "Ochiai-MS Better", "Ochiai Better", "Tie",
                "Both Undetected", "MS Better %", "Ochiai Better %", "Tie %", "Net Advantage"]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        ws3.cell(row=ws3.max_row, column=col).font = header_font
        ws3.cell(row=ws3.max_row, column=col).border = thin_border

    for bt in sorted(type_perf_avg.keys()):
        data = type_perf_avg[bt]
        total = data["Total"]
        ms_better = data["Ochiai-MS Better"]
        oc_better = data["Ochiai Better"]
        tie = data["Tie"]
        undetected = data["Both Undetected"]
        detectable = total - undetected
        ms_pct = (ms_better / detectable * 100) if detectable > 0 else 0
        oc_pct = (oc_better / detectable * 100) if detectable > 0 else 0
        tie_pct = (tie / detectable * 100) if detectable > 0 else 0
        net = ms_better - oc_better
        net_str = f"+{net}" if net > 0 else ("0" if net == 0 else str(net))
        ws3.append([bt, total, blank_if_zero(ms_better), blank_if_zero(oc_better),
                    blank_if_zero(tie), blank_if_zero(undetected),
                    f"{ms_pct:.1f}%", f"{oc_pct:.1f}%", f"{tie_pct:.1f}%", net_str])

    total_all = sum(d["Total"] for d in type_perf_avg.values())
    total_ms = sum(d["Ochiai-MS Better"] for d in type_perf_avg.values())
    total_oc = sum(d["Ochiai Better"] for d in type_perf_avg.values())
    total_tie = sum(d["Tie"] for d in type_perf_avg.values())
    total_und = sum(d["Both Undetected"] for d in type_perf_avg.values())
    net_total = total_ms - total_oc
    ws3.append(["TOTAL", total_all, total_ms, total_oc, total_tie, total_und,
                "", "", "", f"+{net_total}" if net_total > 0 else str(net_total)])
    ws3.cell(row=ws3.max_row, column=1).font = header_font

    # --- Dense Ranking Section ---
    ws3.append([""])
    ws3.append([""])
    ws3.append(["DENSE RANKING"])
    ws3.cell(row=ws3.max_row, column=1).font = title_font
    ws3.append([""])

    type_perf_dense = defaultdict(lambda: {"Ochiai-MS Better": 0, "Ochiai Better": 0, "Tie": 0, "Both Undetected": 0, "Total": 0})
    for rec in bug_records:
        bt = rec["Bug Type"]
        winner = rec["Winner (Dense Ranking)"]
        type_perf_dense[bt][winner] += 1
        type_perf_dense[bt]["Total"] += 1

    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        ws3.cell(row=ws3.max_row, column=col).font = header_font
        ws3.cell(row=ws3.max_row, column=col).border = thin_border

    for bt in sorted(type_perf_dense.keys()):
        data = type_perf_dense[bt]
        total = data["Total"]
        ms_better = data["Ochiai-MS Better"]
        oc_better = data["Ochiai Better"]
        tie = data["Tie"]
        undetected = data["Both Undetected"]
        detectable = total - undetected
        ms_pct = (ms_better / detectable * 100) if detectable > 0 else 0
        oc_pct = (oc_better / detectable * 100) if detectable > 0 else 0
        tie_pct = (tie / detectable * 100) if detectable > 0 else 0
        net = ms_better - oc_better
        net_str = f"+{net}" if net > 0 else ("0" if net == 0 else str(net))
        ws3.append([bt, total, blank_if_zero(ms_better), blank_if_zero(oc_better),
                    blank_if_zero(tie), blank_if_zero(undetected),
                    f"{ms_pct:.1f}%", f"{oc_pct:.1f}%", f"{tie_pct:.1f}%", net_str])

    total_ms_d = sum(d["Ochiai-MS Better"] for d in type_perf_dense.values())
    total_oc_d = sum(d["Ochiai Better"] for d in type_perf_dense.values())
    total_tie_d = sum(d["Tie"] for d in type_perf_dense.values())
    total_und_d = sum(d["Both Undetected"] for d in type_perf_dense.values())
    net_d = total_ms_d - total_oc_d
    ws3.append(["TOTAL", total_all, total_ms_d, total_oc_d, total_tie_d, total_und_d,
                "", "", "", f"+{net_d}" if net_d > 0 else str(net_d)])
    ws3.cell(row=ws3.max_row, column=1).font = header_font

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = max(16, len(headers3[col-1]) + 3)

    # ===== SHEET 4: BY PROJECT =====
    ws4 = wb.create_sheet("By Project")

    ws4.append(["PER-PROJECT BREAKDOWN: Bug types and performance within each project (Average Ranking)."])
    ws4.append(["Shows the distribution of bug types per project and how each algorithm performs."])
    ws4.append([""])

    for project in project_order:
        project_bugs = [r for r in bug_records if r["Project"] == project]
        if not project_bugs:
            continue

        ws4.append([f"{project} ({len(project_bugs)} bugs)"])
        ws4.cell(row=ws4.max_row, column=1).font = title_font
        ws4.append([""])

        proj_headers = ["Bug Type", "Count", "Ochiai-MS Better", "Ochiai Better", "Tie", "Both Undetected"]
        ws4.append(proj_headers)
        for col in range(1, len(proj_headers) + 1):
            ws4.cell(row=ws4.max_row, column=col).font = header_font
            ws4.cell(row=ws4.max_row, column=col).border = thin_border

        proj_type_perf = defaultdict(lambda: {"Ochiai-MS Better": 0, "Ochiai Better": 0, "Tie": 0, "Both Undetected": 0, "Total": 0})
        for rec in project_bugs:
            bt = rec["Bug Type"]
            winner = rec["Winner (Avg Ranking)"]
            proj_type_perf[bt][winner] += 1
            proj_type_perf[bt]["Total"] += 1

        for bt in sorted(proj_type_perf.keys()):
            data = proj_type_perf[bt]
            ws4.append([bt, data["Total"], blank_if_zero(data["Ochiai-MS Better"]),
                       blank_if_zero(data["Ochiai Better"]), blank_if_zero(data["Tie"]),
                       blank_if_zero(data["Both Undetected"])])

        p_ms = sum(1 for r in project_bugs if r["Winner (Avg Ranking)"] == "Ochiai-MS Better")
        p_oc = sum(1 for r in project_bugs if r["Winner (Avg Ranking)"] == "Ochiai Better")
        p_tie = sum(1 for r in project_bugs if r["Winner (Avg Ranking)"] == "Tie")
        p_und = sum(1 for r in project_bugs if r["Winner (Avg Ranking)"] == "Both Undetected")
        ws4.append(["TOTAL", len(project_bugs), blank_if_zero(p_ms), blank_if_zero(p_oc),
                   blank_if_zero(p_tie), blank_if_zero(p_und)])
        ws4.cell(row=ws4.max_row, column=1).font = header_font

        ws4.append([""])
        ws4.append([""])

    for col in range(1, len(proj_headers) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 20

    # ===== SHEET 5: SUMMARY STATISTICS =====
    ws5 = wb.create_sheet("Summary Statistics")

    ws5.append(["SUMMARY STATISTICS FOR PAPER WRITING"])
    ws5.cell(row=1, column=1).font = title_font
    ws5.append(["Key findings formatted for direct inclusion in Chapter 4 Results & Discussion."])
    ws5.append([""])

    # Bug type distribution
    ws5.append(["BUG TYPE DISTRIBUTION"])
    ws5.cell(row=ws5.max_row, column=1).font = title_font
    ws5.append([""])
    ws5.append(["Bug Type", "Count", "Percentage"])
    for col in range(1, 4):
        ws5.cell(row=ws5.max_row, column=col).font = header_font
        ws5.cell(row=ws5.max_row, column=col).border = thin_border

    type_counts = Counter(r["Bug Type"] for r in bug_records)
    for bt, count in type_counts.most_common():
        ws5.append([bt, count, f"{count/len(bug_records)*100:.1f}%"])

    ws5.append([""])
    ws5.append([""])

    # Key findings (avg)
    ws5.append(["KEY FINDINGS (Average Ranking)"])
    ws5.cell(row=ws5.max_row, column=1).font = title_font
    ws5.append([""])
    ws5.append(["Metric", "Value"])
    ws5.cell(row=ws5.max_row, column=1).font = header_font
    ws5.cell(row=ws5.max_row, column=2).font = header_font

    ws5.append(["Total bugs analyzed", len(bug_records)])
    ws5.append(["Ochiai-MS Better (total)", total_ms])
    ws5.append(["Ochiai Better (total)", total_oc])
    ws5.append(["Tied (total)", total_tie])
    ws5.append(["Both Undetected (total)", total_und])
    ws5.append([""])

    # MS advantage types
    ws5.append(["BUG TYPES WHERE OCHIAI-MS HAS NET ADVANTAGE"])
    ws5.cell(row=ws5.max_row, column=1).font = title_font
    ws5.append(["(More bugs where Ochiai-MS outperforms than underperforms)"])
    ws5.append([""])
    adv_headers = ["Bug Type", "MS Better", "Ochiai Better", "Tie", "Net Advantage", "Total"]
    ws5.append(adv_headers)
    for col in range(1, len(adv_headers) + 1):
        ws5.cell(row=ws5.max_row, column=col).font = header_font
        ws5.cell(row=ws5.max_row, column=col).border = thin_border

    for bt in sorted(type_perf_avg.keys()):
        data = type_perf_avg[bt]
        net = data["Ochiai-MS Better"] - data["Ochiai Better"]
        if net > 0:
            ws5.append([bt, data["Ochiai-MS Better"], data["Ochiai Better"],
                       data["Tie"], f"+{net}", data["Total"]])

    ws5.append([""])

    # Ochiai advantage types
    ws5.append(["BUG TYPES WHERE OCHIAI HAS NET ADVANTAGE"])
    ws5.cell(row=ws5.max_row, column=1).font = title_font
    ws5.append(["(More bugs where Standard Ochiai outperforms Ochiai-MS)"])
    ws5.append([""])
    ws5.append(adv_headers)
    for col in range(1, len(adv_headers) + 1):
        ws5.cell(row=ws5.max_row, column=col).font = header_font
        ws5.cell(row=ws5.max_row, column=col).border = thin_border

    for bt in sorted(type_perf_avg.keys()):
        data = type_perf_avg[bt]
        net = data["Ochiai-MS Better"] - data["Ochiai Better"]
        if net < 0:
            ws5.append([bt, data["Ochiai-MS Better"], data["Ochiai Better"],
                       data["Tie"], str(net), data["Total"]])

    ws5.append([""])

    # Neutral types
    ws5.append(["BUG TYPES WITH NO CLEAR WINNER (Neutral)"])
    ws5.cell(row=ws5.max_row, column=1).font = title_font
    ws5.append(["(Equal performance between both algorithms)"])
    ws5.append([""])
    ws5.append(adv_headers)
    for col in range(1, len(adv_headers) + 1):
        ws5.cell(row=ws5.max_row, column=col).font = header_font
        ws5.cell(row=ws5.max_row, column=col).border = thin_border

    for bt in sorted(type_perf_avg.keys()):
        data = type_perf_avg[bt]
        net = data["Ochiai-MS Better"] - data["Ochiai Better"]
        if net == 0:
            ws5.append([bt, data["Ochiai-MS Better"], data["Ochiai Better"],
                       data["Tie"], "0", data["Total"]])

    ws5.column_dimensions['A'].width = 48
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 10
    ws5.column_dimensions['E'].width = 16
    ws5.column_dimensions['F'].width = 10

    # --- SAVE ---
    output_path = os.path.join(base_dir, "Bug_Type_Analysis.xlsx")
    wb.save(output_path)
    print(f"\nDone! Report saved to: {output_path}")
    print(f"\n-- QUICK SUMMARY (Average Ranking) --")
    print(f"  Total bugs: {len(bug_records)}")
    print(f"  Ochiai-MS Better: {total_ms}")
    print(f"  Ochiai Better:    {total_oc}")
    print(f"  Tied:             {total_tie}")
    print(f"  Both Undetected:  {total_und}")
    print(f"\n-- Bug Type Distribution --")
    for bt, count in type_counts.most_common():
        print(f"  {bt:25s} {count:3d} ({count/len(bug_records)*100:.1f}%)")


if __name__ == "__main__":
    main()
